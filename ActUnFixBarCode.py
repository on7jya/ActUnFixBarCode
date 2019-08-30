import pathlib
import random
import time
import datetime

import paramiko

from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QFileDialog
import openpyxl
import os
from loggers import Loggers
from properties import *


class Ui_MainWindow(object):
    def __init__(self):
        super().__init__()
        
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1066, 457)
        MainWindow.setMinimumSize(QtCore.QSize(777, 333))
        MainWindow.setMaximumSize(QtCore.QSize(777, 333))
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        self.ButtonSelectFile = QtWidgets.QPushButton(self.centralwidget)
        self.ButtonSelectFile.setGeometry(QtCore.QRect(10, 10, 150, 31))
        self.ButtonSelectFile.setObjectName("ButtonSelectFile")

        self.ButtonGO = QtWidgets.QPushButton(self.centralwidget)
        self.ButtonGO.setGeometry(QtCore.QRect(10, 50, 150, 31))
        self.ButtonGO.setObjectName("ButtonGO")

        self.linePathtoFile = QtWidgets.QLineEdit(self.centralwidget)
        self.linePathtoFile.setGeometry(QtCore.QRect(170, 20, 430, 21))
        self.linePathtoFile.setObjectName("linePathtoFile")

        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(170, 0, 81, 20))
        self.label.setObjectName("label")

        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(10, 100, 750, 200))
        self.textEdit.setObjectName("textEdit")

        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "ActFixBarCode"))
        self.ButtonSelectFile.setText(_translate("MainWindow", "Выбрать файл"))
        self.ButtonGO.setText(_translate("MainWindow", "Запустить"))
        self.label.setText(_translate("MainWindow", "Путь до файла:"))
        self.ButtonSelectFile.clicked.connect(self.get_path_to_file)
        self.ButtonGO.clicked.connect(self.main)

    def get_path_to_file(self):
        fileName, _ = QFileDialog.getOpenFileName(None, "Открыть файл", "","Excel Files (*.xls*)")
        if fileName:
            self.linePathtoFile.setText(fileName)

    def read_workbook(self):
        path_to_file = self.linePathtoFile.text()
        try:
            self.wb = openpyxl.load_workbook(path_to_file)
            self.ws = self.wb.active
            return True
        except:
            self.statusbar.setStyleSheet("color: rgb(255, 0, 0)")
            self.statusbar.showMessage('Не удалось загрузить файл Excel', msecs=3000)
            return False

    # def generate_insert_query_text(self, args_list):
    #     def generate_text(args):
    #         string = f"INSERT INTO BARCODE_FROM_SAP(store_adress_number, plu, e_fb, barcode) VALUES ('{args[0]}','{args[1]}','{args[2]}','{args[3]}');"
    #         return string
    #     query_text = list(map(generate_text, args_list))
    #     return ' '.join(query_text)


    # def create_plu_counts_dict(self):
    #     try:
    #         plu_list = []
    #         for i in range(1, self.ws.max_row + 1):
    #             if not str(self.ws.cell(row=i, column=2).value).isalpha():
    #                 plu_list.append(str(self.ws.cell(row=i, column=2).value))
    #         plu_short_list = set(plu_list)
    #         for e in plu_short_list:
    #             self.plu_counts_dict.update({e: plu_list.count(e)})
    #         return True
    #     except Exception as e:
    #         print('function create_plu_counts_dict', str(e))
    #         return False


    def parse_excel_to_list(self):
        try:
            if len(self.ws.cell(row=1, column=4).value) < 50:
                start = 2
            else:
                start = 1
            _list = []
            for i in range(start, self.ws.max_row + 1):
                _list.append([str(self.ws.cell(row=i, column=1).value),
                              str(self.ws.cell(row=i, column=2).value),
                              str(self.ws.cell(row=i, column=3).value),
                              str(self.ws.cell(row=i, column=4).value)])
            print(_list)
            return _list
        except Exception as e:
            logger.error(e)
            self.statusbar.setStyleSheet("color: rgb(255, 0, 0)")
            self.statusbar.showMessage(e, msecs=3000)
            return False

    def main(self):
        if os.path.isfile(self.linePathtoFile.text()):
            if self.read_workbook():
                path_to_file = self.linePathtoFile.text()
                list_fix = self.parse_excel_to_list()
                self.go(list_fix)
        else:
            self.statusbar.setStyleSheet("color: rgb(255, 0, 0)")
            self.statusbar.showMessage('Не выбран файл', msecs = 3000)

    def go(self, _list):
        for i in range(len(_list)):
            print(i)
            self.textEdit.append(str(i+1))
            try:
                sap_id = ''
                sap_id = _list[i] [0]
                sap_id = sap_id.strip()
                fsrar_id = ''
                fsrar_id = _list[i][1]
                fsrar_id = fsrar_id.strip()
                efb = ''
                efb = _list[i][2]
                efb = efb.strip()
                barcode = ''
                barcode = _list[i][3]
                barcode = barcode.strip()

                print('sap_id:' + sap_id )
                self.textEdit.append('sap_id:' + sap_id )
                logger.info(sap_id)
                print('fsrar_id:' + fsrar_id)
                self.textEdit.append('fsrar_id:' + fsrar_id )
                print('efb:' + efb)
                self.textEdit.append('efb:' + efb )
                print('Barcode:' + barcode)
                self.textEdit.append('barcode:' + barcode )
                logger.info(barcode)

                # Создание xml и наполнение данными
                name_file = f"""ActUnFixBarCode_{sap_id}_{barcode}.xml""".format(sap_id,barcode)

                # pathlib.Path(f"./{sap_id}".format(sap_id)).mkdir(parents=True, exist_ok=True)
                # logger.info('10')
                # file = open(f"./{sap_id}/".format(sap_id) + name_file, 'w', encoding='utf-8')

                pathlib.Path(f"./result").mkdir(parents=True, exist_ok=True)
                file = open(r"./result/" + name_file, 'w', encoding='utf-8')

                # file.write(f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
                # <Cheque datetime="{8}" number="{3}" shift="{4}" kassa="{2}" name="{1}" address="{0}" kpp="{9}" inn="{10}">
                # <Bottle ean="{5}" barcode="{6}" price="{7}"/></Cheque>""".format(Address,Name,Kassa,Number,Shift,Ean,Barcode,Price,Date,Kpp,Inn))

                number_act = random.randint(1,99999999)
                date_act = datetime.datetime.now().strftime("%Y-%m-%d")
                file.write(f"""<?xml version="1.0" encoding="UTF-8"?>
            <ns:Documents Version="1.0"
            xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
            xmlns:ns= "http://fsrar.ru/WEGAIS/WB_DOC_SINGLE_01"
            xmlns:pref="http://fsrar.ru/WEGAIS/ProductRef_v2"
            xmlns:awr="http://fsrar.ru/WEGAIS/ActUnFixBarCode"
            xmlns:ce="http://fsrar.ru/WEGAIS/CommonV3">
            	<ns:Owner>
            		<ns:FSRAR_ID>{fsrar_id}</ns:FSRAR_ID>
            	</ns:Owner>
            	<ns:Document>
            		<ns:ActUnFixBarCode>
            			<awr:Identity>1</awr:Identity>
            			<awr:Header>
            				<awr:Number>{number_act}</awr:Number>
            				<awr:ActDate>{date_act}</awr:ActDate>
            				<awr:Note></awr:Note>
            			</awr:Header>
            			<awr:Content>
            				<awr:Position>
            					<awr:Identity>1</awr:Identity>
            					<awr:Inform2RegId>{efb}</awr:Inform2RegId>
            					<awr:MarkInfo>
            						<ce:amc>{barcode}</ce:amc>
            					</awr:MarkInfo>
            				</awr:Position>
            			</awr:Content>
            		</ns:ActUnFixBarCode>
            	</ns:Document>
            </ns:Documents>""".format(fsrar_id, number_act, efb, barcode ))
                file.close()
                time.sleep(2)
                self.textEdit.append('xml создан')
                res = 1
                self.sender(sap_id, name_file)

                #return res

            except Exception as e:
                logger.error(e)
                self.statusbar.setStyleSheet("color: rgb(255, 0, 0)")
                self.statusbar.showMessage(e, msecs=3000)

    def sender(self, sap_id, name_file):
        pos_client = paramiko.SSHClient()
        pos_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            pos_client.connect(hostname='BO-' + sap_id, username=login_server_gk, password=pass_server_gk, port=port_server_gk, timeout=5)
            sftp = pos_client.open_sftp()
            #sftp.put(r'ActFixBarCode_307Z_102400004576501018001Z6UJICFTHHELB3KPGUY4AH5QMI3CP5BLIZMQBURNHQBRHSITGI3OBPRQSYCO2G6ARSNY6BST6JSSOLSCX7GTTWKPJR5RZAN4XJCSQYQO6ZO42XA3WXOMACBKVQVNDXHFI.xml', r'/usr/local/gkretail/ActFixBarCode_307Z_102400004576501018001Z6UJICFTHHELB3KPGUY4AH5QMI3CP5BLIZMQBURNHQBRHSITGI3OBPRQSYCO2G6ARSNY6BST6JSSOLSCX7GTTWKPJR5RZAN4XJCSQYQO6ZO42XA3WXOMACBKVQVNDXHFI.xml')
            sftp.put(f"""./result/{name_file}""".format(name_file), f"""/usr/local/gkretail/{name_file}""".format(name_file))

            time.sleep(2)
            sftp.close()

            stdin, stdout, stderr = pos_client.exec_command(f"find /usr/local/gkretail -name {name_file}".format(name_file) )
            out = stdout.read().strip().decode('utf-8')

            if f"{name_file}".format(name_file) in out:
                self.textEdit.append('Файл загружен на сервер')

                stdin, stdout, stderr = pos_client.exec_command(
                    f'curl -F "xml_file=@/usr/local/gkretail/{name_file}" http://BO-'.format(name_file) + sap_id + ':8195/opt/in/ActUnFixBarCode')  # Посыл xml на УТМ
                out = stdout.read().strip().decode('utf-8')
                self.textEdit.append('Ответ от УТМ:')
                self.textEdit.append(out)
                self.textEdit.append('---------------------------------------------------------------------------------------------------------------------------------')
                logger.info(out)
                print(out)

            else:
                self.textEdit.append(out)


            pos_client.close()
        except:
            self.statusbar.setStyleSheet(
                "color: rgb(255, 0, 0)")  # делаем цвет текста статусбара (нижней строки приложения для вывода статусов) красным
            self.statusbar.showMessage('Ошибка подключения по SSH',
                                       msecs=1000)  # в статусбар пишем сообщение об ошибке, ошибка будет отображаться 10000 милисекунд
            self.textEdit.append('Произошла ошибка во время операции')


if __name__ == "__main__":
    import sys

    logger = Loggers.logging(1, "MainWindow", "DEBUG")

    QtWidgets.QApplication.setStyle('fusion')
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    #ui.start_check()
    MainWindow.show()
    sys.exit(app.exec_())

